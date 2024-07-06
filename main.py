from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import logging
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def american_to_decimal(american_odds):
    if american_odds == "" or american_odds == "-" or american_odds.startswith("U") or american_odds.startswith("O"):
        return 0
    if american_odds.startswith('+'):
        return round(int(american_odds) / 100 + 1, 2)
    elif american_odds.startswith('-'):
        return round(100 / abs(int(american_odds)) + 1, 2)

def scrape_mlb_games_and_odds():
    url = "https://sports.yahoo.com/mlb/odds/"
    driver = setup_driver()
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "odds-table-v2"))
        )
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')
        
        odds_table = soup.find('div', class_='odds-table-v2')        
        games = []
        
        game_divs = odds_table.find_all('div', class_='PREGAME sixpack') or \
                    odds_table.find_all('div', class_='sixpack') or \
                    odds_table.find_all('div', class_='Bgc(#fff)')
        
        for game_div in game_divs:
            team_spans = game_div.find_all('span', class_='Fw(600)') or \
                         game_div.find_all('span', class_='Ell D(ib) Va(m)')
            
            moneyline_buttons = game_div.find_all('button', class_='YahooSans-Semibold') or \
                                game_div.find_all('button', class_='Bdrs(2px) Fw(600)')
            
            if len(team_spans) >= 2 and len(moneyline_buttons) >= 2:
                away_team = team_spans[0].text.strip()
                home_team = team_spans[1].text.strip()
                away_moneyline = moneyline_buttons[0].text.strip()
                away_spread = moneyline_buttons[1].text.strip()
                home_moneyline = moneyline_buttons[3].text.strip()
                home_spread = moneyline_buttons[4].text.strip()
                total_runs_under = moneyline_buttons[5].text.strip()
                total_runs_over = moneyline_buttons[2].text.strip()
                
                game_info = {
                    'away_team': away_team,
                    'home_team': home_team,
                    'away_moneyline': away_moneyline,
                    'home_moneyline': home_moneyline,
                    'away_spread': away_spread,
                    'home_spread': home_spread,
                    'total_runs_under': total_runs_under,
                    'total_runs_over': total_runs_over
                }
                games.append(game_info)
        return games
    finally:
        driver.quit()

def save_to_excel(games):    
    wb = Workbook()
    ws = wb.active
    ws.title = "MLB Odds"
    headers = ['Date/Time', 'Team', 'Money Line', 'Run Spread', 'Total Runs', "", 'Home power', 'val', "", 'Home ML', 'val', "", 'Guest ML', 'val', "", 'Favorite Team', 'val', "", 'Underdog Team', 'val', "", '+1.5 Spread ', 'val', "", '-1.5 Spread ', 'val']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row = 2
    for game in games:
        ws.cell(row=row, column=1, value=str(datetime.today().month) + "/" + str(datetime.today().day) + ' ' + str(datetime.today().strftime("%A")))  
        ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
        
        game['away_team'] = game['away_team'].split(" ")[-1] if game['away_team'].split(" ")[-1] != "Sox" else "".join(game['away_team'].split(" ")[-2:])
        game["away_moneyline"] = (str(american_to_decimal(game['away_moneyline']))) + "   " +  "(" + game['away_moneyline'] + ")"
        game['away_spread'] = (game['away_spread'][:4] + " "  + str(american_to_decimal(game['away_spread'][4:]))) + "   " +  "(" + game['away_spread'][4:] + ")"

        game['total_runs_under']= (game['total_runs_under'][:-4] + " "  + str(american_to_decimal(game['total_runs_under'][-4:]))) + "   " +  "(" + game['total_runs_under'][-4:] + ")"

        game['home_team'] = game['home_team'].split(" ")[-1] if game['home_team'].split(" ")[-1] != "Sox" else "".join(game['home_team'].split(" ")[-2:])
        game['home_moneyline'] = (str(american_to_decimal(game['home_moneyline']))) + "   " +  "(" + game['home_moneyline'] + ")"
        game['home_spread'] = (game['home_spread'][:4] + " "  + str(american_to_decimal(game['home_spread'][4:]))) + "   " +  "(" + game['home_spread'][4:] + ")"
        game['total_runs_over']= (game['total_runs_over'][:-4] + " "  + str(american_to_decimal(game['total_runs_over'][-4:]))) + "   " +  "(" + game['total_runs_over'][-4:] + ")"
        
        # Away team
        ws.cell(row=row, column=2, value=game['away_team'])
        ws.cell(row=row, column=3, value=game['away_moneyline'])
        ws.cell(row=row, column=4, value=game['away_spread'][1:] if game['away_spread'].startswith('+') else game['away_spread'])
        ws.cell(row=row, column=5, value= game['total_runs_under'])
        
        # Home team
        row += 1
        ws.cell(row=row, column=2, value=game['home_team'])
        ws.cell(row=row, column=3, value=game['home_moneyline'])
        ws.cell(row=row, column=4, value=game['home_spread'][1:] if game['home_spread'].startswith('+') else game['home_spread'])
        ws.cell(row=row, column=5, value=game['total_runs_over'])
        
        # Home power
        if game['home_spread'][:4] == '-1.5':
            home_bet_team = f"{game['home_team']} (-1.5)"
            home_bet_decimal = float(game['home_spread'][4:-6]) * 100 - 100 if game['home_spread'][4:-6] != " " else 0
        else:
            home_bet_team = f"{game['home_team']} (ML)"
            home_bet_decimal = float(game['home_moneyline'].split()[0]) * 100 - 100

        ws.cell(row=row, column=7, value=home_bet_team)
        ws.cell(row=row, column=8, value=home_bet_decimal)
        ws.cell(row=row, column=8).number_format = '0.00'
        
        # Home ML Bet
        home_ml_decimal = float(game['home_moneyline'].split()[0]) * 100 - 100
        ws.cell(row=row, column=10, value=game['home_team'])
        ws.cell(row=row, column=11, value=home_ml_decimal)
        ws.cell(row=row, column=11).number_format = '0.00'
        
        row += 1  

        # Away ML Bet
        away_ml_decimal = float(game['away_moneyline'].split()[0]) * 100 - 100
        ws.cell(row=row-1, column=13, value=game['away_team'])
        ws.cell(row=row-1, column=14, value=away_ml_decimal)
        ws.cell(row=row-1, column=14).number_format = '0.00'

        # Favorite and Underdog ML Bet
        if home_ml_decimal <= away_ml_decimal:
            fav_team = game['home_team']
            fav_ml_decimal = home_ml_decimal
            dog_team = game['away_team']
            dog_ml_decimal = away_ml_decimal
        else:
            fav_team = game['away_team']
            fav_ml_decimal = away_ml_decimal
            dog_team = game['home_team']
            dog_ml_decimal = home_ml_decimal
        
        ws.cell(row=row-1, column=16, value=fav_team)
        ws.cell(row=row-1, column=17, value=fav_ml_decimal)
        ws.cell(row=row-1, column=17).number_format = '0.00'

        ws.cell(row=row-1, column=19, value=dog_team)
        ws.cell(row=row-1, column=20, value=dog_ml_decimal)
        ws.cell(row=row-1, column=20).number_format = '0.00'

        # +1.5 Spread Bet
        if game['away_spread'][:4] == '+1.5':
            spread_team = game['away_team']
            spread_decimal = float(game['away_spread'][4:-6]) * 100 - 100 if game['away_spread'][4:-6] != " " else 0
        elif game['home_spread'][:4] == '+1.5':
            spread_team = game['home_team']
            spread_decimal = float(game['home_spread'][4:-6]) * 100 - 100 if game['home_spread'][4:-6] != " " else 0
        else:
            spread_team = "N/A"
            spread_decimal = 0

        ws.cell(row=row-1, column=22, value=spread_team)
        ws.cell(row=row-1, column=23, value=spread_decimal)
        ws.cell(row=row-1, column=23).number_format = '0.00'

        # -1.5 Spread Bet
        if game['away_spread'][:4] == '-1.5':
            spread_team_minus = game['away_team']
            spread_decimal_minus = float(game['away_spread'][4:-6]) * 100 - 100 if game['away_spread'][4:-6] != " " else 0
        elif game['home_spread'][:4] == '-1.5':
            spread_team_minus = game['home_team']
            spread_decimal_minus = float(game['home_spread'][4:-6]) * 100 - 100 if game['home_spread'][4:-6] != " " else 0
        else:
            spread_team_minus = "N/A"
            spread_decimal_minus = 0

        ws.cell(row=row-1, column=25, value=spread_team_minus)
        ws.cell(row=row-1, column=26, value=spread_decimal_minus)
        ws.cell(row=row-1, column=26).number_format = '0.00'

    for col in range(1, 27):
        ws.column_dimensions[get_column_letter(col)].width = 12

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"mlb_odds_{timestamp}.xlsx"
    wb.save(filename)
    logging.info(f"Data saved to {filename}")

if __name__ == "__main__":
    games = scrape_mlb_games_and_odds()
    if games:
        save_to_excel(games)